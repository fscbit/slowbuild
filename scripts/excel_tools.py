# Excel工具：批量xls转xlsx + 按列找重复行
# 两步：1. 当前目录所有.xls → .xlsx  2. 对转换后的文件按G列查重导出
import os
import openpyxl

print("=" * 40)
print("步骤1: 批量转换 .xls → .xlsx")
print("=" * 40)

# 获取所有xls文件
file_list = [f for f in os.listdir('.') if f.endswith('.xls') and not f.startswith('~')]

if not file_list:
    print("当前目录没有找到 .xls 文件")
else:
    for file_name in file_list:
        try:
            import xlrd
            xls_file = os.path.join('.', file_name)
            workbook = xlrd.open_workbook(xls_file)
            sheet_names = workbook.sheet_names()
            
            xlsx_file = os.path.splitext(file_name)[0] + '.xlsx'
            workbook_new = openpyxl.Workbook()
            
            for idx, sheet_name in enumerate(sheet_names):
                if idx == 0:
                    worksheet_new = workbook_new.active
                    worksheet_new.title = sheet_name
                else:
                    worksheet_new = workbook_new.create_sheet(title=sheet_name)
                worksheet = workbook.sheet_by_name(sheet_name)
                for row in range(worksheet.nrows):
                    for col in range(worksheet.ncols):
                        cell_value = worksheet.cell_value(row, col)
                        worksheet_new.cell(row=row + 1, column=col + 1, value=cell_value)
            
            workbook_new.save(xlsx_file)
            os.remove(xls_file)
            print(f"✅ {file_name} → {xlsx_file}")
        except Exception as e:
            print(f"❌ {file_name} 转换失败: {e}")

print("\n" + "=" * 40)
print("步骤2: 按第7列(G列)查重")
print("=" * 40)

# 自动找第一个xlsx文件查重
xlsx_files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('~')]

if not xlsx_files:
    print("没有找到 .xlsx 文件来查重")
else:
    input_file = xlsx_files[0]
    print(f"分析文件: {input_file}")
    
    from collections import defaultdict
    wb1 = openpyxl.load_workbook(input_file)
    ws1 = wb1.active
    
    # 按G列(索引6)分组
    duplicates = defaultdict(list)
    for row in ws1.iter_rows(min_row=2, values_only=True):
        g_value = row[6] if len(row) > 6 else None
        if g_value is not None and g_value != '':
            duplicates[g_value].append(row)
    
    # 创建结果工作簿
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    
    # 写入表头
    headers = [cell.value for cell in ws1[1]] + ['重复次数']
    ws2.append(headers)
    
    dup_count = 0
    for g_val, rows in duplicates.items():
        if len(rows) <= 1:
            continue
        sorted_rows = sorted(rows, key=lambda x: x[6] if len(x) > 6 else '')
        for row in sorted_rows:
            ws2.append(list(row) + [len(rows) - 1])
        dup_count += 1
        print(f"  G列='{g_val}' 重复 {len(rows)} 次")
    
    output_file = '重复行结果.xlsx'
    wb2.save(output_file)
    print(f"\n✅ 找到 {dup_count} 组重复，已导出: {output_file}")

print("\n全部完成！")
input("按回车退出...")
